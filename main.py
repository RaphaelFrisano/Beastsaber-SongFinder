#OS and Time import for Data formatting etc.
import time
import os

# Openpyxel Import for Excel
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

# Selenium Imports for Browser
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import DesiredCapabilities
from selenium.webdriver import Firefox

# Spotipy import for Spotify
import spotipy
from spotipy import util
from spotipy.oauth2 import SpotifyClientCredentials



if __name__ == '__main__':

    def get_playlist_tracks(sp, sp_username, playlist_id):
        # Gets all Tracks from playlists even if they have over 100 Songs
        results = sp.user_playlist_tracks(sp_username,playlist_id)
        tracks = results['items']
        while results['next']:
            results = sp.next(results)
            tracks.extend(results['items'])
        return tracks

    def format_output_playlists_song():
        # Output all Songs to console
        # ! Requires get_playlist_tracks() function
        playlist_id = input("Your playlist id: ")
        trackslist = get_playlist_tracks(sp, sp_username, playlist_id)

        for item in trackslist:
            songname = item['track']['name']
            songid = item['track']['id']
            mainartistname = item['track']['artists'][0]['name']
            mainartistid = item['track']['artists'][0]['id']
            print(songname, mainartistname, sep=' - ')

    def setup_spotify(client_id, client_secret, sp_username, scope, redirect_uri):
        # Sets up a Spotify Client for User with generated Client + Secret
        client_credentials_manager = SpotifyClientCredentials(client_id=client_id, client_secret=client_secret)
        sp = spotipy.Spotify(client_credentials_manager=client_credentials_manager)
        token = util.prompt_for_user_token(sp_username, scope, client_id, client_secret, redirect_uri)
        if token:
            sp = spotipy.Spotify(auth=token)
            return sp
        else:
            print("Can't get token for ", sp_username)
            return ("Can't get token for " + sp_username)

    def main():
        # <----------------------------------->
        # <---| Set up Spotify Connection |--->
        # <----------------------------------->
        # <-| Variable Definition |->
        client_id = 'd5ec1915e2b3452f87cd1f224551a935'
        f = open("secret.txt", "r")
        client_secret = str(f.read())
        sp_username = '16r49f73ryoeuabwxqwgpimzs'
        scope = 'user-library-read playlist-modify-public playlist-read-private'
        redirect_uri='http://localhost:8888/callback'
        # <-| Generate Spotify Object and save into variable "sp" |->
        sp = setup_spotify(client_id, client_secret, sp_username, scope, redirect_uri)

        #Connect to playlist
        playlist_id = input("Your playlist id: ")
        trackslist = get_playlist_tracks(sp, sp_username, playlist_id)

        #Open/Create Excels
        allxlxwb = load_workbook(os.getcwd() + "/All_Found_Songs.xlsx")
        allxlx = allxlxwb['Tabelle1']
        
        newxlxwb = Workbook()
        newxlxwb.save(os.getcwd() + "/New_Found_Songs.xlsx")
        newxlxwb = load_workbook(os.getcwd() + "/New_Found_Songs.xlsx")
        newxlx = newxlxwb['Sheet']
        newxlx.cell(row = 1, column = 1).value = "Artist"
        newxlx.cell(row = 1, column = 2).value = "Title"

        #Create Browser and open Beastsaber on site to search
        browser = webdriver.Firefox(os.getcwd()) # Workingdir -> Finds gecko driver
        browser.get('https://bsaber.com/?s=')
        WebDriverWait(browser,10000).until(EC.visibility_of_element_located((By.TAG_NAME,'body')))

        # Get songs
        excelStartFile = open(os.getcwd() + '/startline.txt', 'r+')
        row = excelStartFile.read()
        row = [int(s) for s in row.split() if s.isdigit()] # Get only numbers out of File
        row = row[0] # Get first number
        for songdict in trackslist:
            # Get song Info
            songtitle = songdict['track']['name']
            artist = songdict['track']['artists'][0]['name']

            print("=======================|>")
            print("Title - " + songtitle)
            print("Artist - " + artist)

            # Check if song already got found on another date
            songfound = False
            i = 1
            while True:
                i = i + 1
                cell = allxlx.cell(row = i, column = 2).value
                if cell == songtitle:
                    songfound = True
                if cell == None:
                    break

            if songfound:
                print("Already found!")
                print("=======================|>")
                print("")
                continue # Song already got found, go to next in loop

            # Check if Song exists
            browser.find_element_by_xpath("/html/body/div[1]/div/div[2]/div/aside/div[1]/div/div[1]/form/fieldset/input[1]").click()
            browser.find_element_by_xpath("/html/body/div[1]/div/div[2]/div/aside/div[1]/div/div[1]/form/fieldset/input[1]").clear()
            browser.find_element_by_xpath("/html/body/div[1]/div/div[2]/div/aside/div[1]/div/div[1]/form/fieldset/input[1]").send_keys(artist + " " + songtitle)
            browser.find_element_by_xpath("/html/body/div[1]/div/div[2]/div/aside/div[1]/div/div[1]/form/fieldset/input[2]").click()

            try:
                browser.find_element_by_partial_link_text(songtitle)
            except NoSuchElementException:
                print("Doesen't exist!")
                print("=======================|>")
                print("")
                continue
            
            print("Song exists!")
            print("=======================|>")
            print("")

            # Add songs to Excel
            allxlx.cell(row = row, column = 1).value = str(artist)
            allxlx.cell(row = row, column = 2).value = str(songtitle)
            allxlx.cell(row = row, column = 3).value = str(time.asctime())

            newxlx.cell(row = row, column = 1).value = str(artist)
            newxlx.cell(row = row, column = 2).value = str(songtitle)

            row = row + 1            

        # End
        browser.quit()
        browser = ""
        excelStartFile.truncate(0)
        excelStartFile.write(str(row))
        excelStartFile.close()
        allxlxwb.save(os.getcwd() + "/All_Found_Songs.xlsx")
        newxlxwb.save(os.getcwd() + "/New_Found_Songs.xlsx")
        input("ENTER in CMD to end: ")

# <---------------------->
# <---| Start Script |--->
# <---------------------->
main()